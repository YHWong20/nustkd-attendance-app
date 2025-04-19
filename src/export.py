"""
Attendance Excel export utilities
"""

from datetime import datetime, timezone, timedelta
import logging
import openpyxl
from src.gdrive import download_file, upload_file
from src.telegram import send_message

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)8s: %(message)s"
)

# Row and column numbers of interest in excel template
DATE_ROW = 8
LEFT_DATE_LIMIT = 5
RIGHT_DATE_LIMIT = 14
NAME_START_ROW = 10
NAME_COL = 2
MAX_NAME_ROW = 170


def get_target_column(ws, export_date):
    """
    Get target column in excel sheet for given export date.

    Args:
        export_date (int): Day to be exported.

    Returns:
        int: Target column number if found, else None.
    """
    for col in range(LEFT_DATE_LIMIT, RIGHT_DATE_LIMIT):
        if int(ws.cell(row=DATE_ROW, column=col).value) == export_date:
            logging.info("Target column for date %s is column %s.", export_date, col)
            return col

    logging.warning("Date %s not found in the sheet!", export_date)
    return None


def zero_missing_members(ws, col_no):
    """
    Fill empty cells in column with 0.

    Args:
        col_no (in): Target column.
    """
    for row in range(NAME_START_ROW + 1, MAX_NAME_ROW):
        cell_value = ws.cell(row=row, column=col_no).value

        if not cell_value or int(cell_value) != 1:
            ws.cell(row=row, column=col_no, value=0)

    logging.info("Zeroed empty cells.")
    return ws


def insert_entries(export_date, entries):
    """
    Insert name entries into Excel sheet.

    Args:
        export_date (int): Day to be exported.
        entries (list): List of Member entries.
    """
    # Load template from google drive
    template = download_file()

    try:
        wb = openpyxl.load_workbook(template)
        ws = wb["Active Members"]
        logging.info("Workbook loaded.")
    except Exception as e:
        logging.error("Workbook could not be loaded. Error: %s", e)
        raise e

    target_date = int(export_date)
    target_column = get_target_column(ws, target_date)

    if not target_column:
        logging.warning(
            "Target column %s not found in Attendance sheet.",
            target_column
        )
        return

    attendance_names = list(map(lambda x: x["name"], entries))
    duplicate_names = []
    missing_names = []

    for name in attendance_names:
        matching_rows = []

        for row in range(NAME_START_ROW + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=NAME_COL).value
            if isinstance(cell_value, str) and name.lower() in cell_value.lower():
                matching_rows.append(row)

        if len(matching_rows) == 1:
            row_to_update = matching_rows[0]
            ws.cell(row=row_to_update, column=target_column, value=1)
            logging.info(
                "Marked %s as present on %s (Row %s).",
                name,
                target_date,
                row_to_update
            )
        elif len(matching_rows) > 1:
            # Duplicate name entry
            # TODO: possible utilise member status to de-duplicate names
            logging.warning(
                "Skipping %s, Found multiple matches: %s",
                name,
                matching_rows
            )

            duplicate_names.append(name)
        else:
            # Name not found
            logging.warning("%s not found.", name)
            missing_names.append(name)

    ws = zero_missing_members(ws, target_column)

    # Send telegram status message
    sgt = timezone(timedelta(hours=8))
    month = datetime.now(sgt).month
    day_month = f"{export_date}/{month}"
    send_message(day_month, duplicates=duplicate_names, missing=missing_names)

    try:
        # Save and upload excel sheet to google drive
        wb.save(template)
        wb.close()
        logging.info("Workbook saved.")
    except Exception as e:
        logging.error("Workbook could not be saved. Error: %s", e)
        raise e

    upload_file()
